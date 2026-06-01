"""Manual scraping workbench API."""
from __future__ import annotations

import asyncio
import io
import json
import logging
import re
from collections import deque
from datetime import datetime, timezone
from urllib.parse import parse_qsl, unquote, urlencode, urljoin, urlparse, urlunparse

import httpx
from bs4 import BeautifulSoup
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openai import AsyncOpenAI
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from backend.ai.extractor import parse_json_safely
from backend.config import settings
from backend.database.connection import _get_engine
from backend.database import crud
from backend.discovery.apify_client import discover_agencies_sync
from backend.scraper.engine import ScraperEngine
from backend.scraper.reference_sniffer import collect_reference_tokens_from_html

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/workbench", tags=["workbench"])

_engine_singleton = ScraperEngine()


def _urls_containing_reference(ref: str, urls: list[str], *, limit: int = 80) -> list[str]:
    """Internal links whose path/query contains this reference substring (case-insensitive)."""
    if not ref or not urls:
        return []
    rlow = ref.lower()
    hits = [u for u in urls if rlow in u.lower()]
    return hits[:limit]


COMPREHENSIVE_EXTRACT_PROMPT = """
You are an expert web data extractor.
Extract EVERY SINGLE piece of information
visible on this webpage.

This could be a real estate property page,
an about page, a contact page, or any other page.

INSTRUCTIONS:
1. Look at ALL text, numbers, labels on the page
2. Extract every data point you can find
3. Create appropriate field names for each
4. Return as flat JSON object (no nested objects
   except arrays)
5. Field names: lowercase with underscores
   e.g. "price_eur", "floor_number", "agent_name"

For real estate pages, look for:
- reference_number (REF:, Property ID, or UUID in URL path e.g. /listings/bd199e03-16e1-4dad-ac3e-658adced81ad)
- title, subtitle
- price (number only), price_text (formatted)
- currency
- property_type (apartment/villa/etc)
- category (sale/rent)
- status (on market/sold/etc)
- badge (sole agency/new/etc)
- bedrooms (int)
- bathrooms (int)
- internal_sqm (float)
- external_sqm (float)
- total_sqm (float)
- plot_sqm (float)
- floor_number
- total_floors
- year_built
- furnished (yes/no/part)
- condition (new/good/etc)
- locality
- town
- region
- country
- full_address
- latitude (from map if present)
- longitude (from map if present)
- description (FULL text, not truncated)
- features (array of strings)
- amenities (array of strings)
- energy_rating
- permit_number
- agent_name
- agent_phone
- agent_email
- agency_name
- owner_name, owner_phone, owner_email (property vendor / landlord / seller — NOT the estate agent)
- price_per_night (numeric nightly rent if shown, e.g. holiday lets)
- price_per_month (numeric monthly rent if shown)
- has_wifi (bool) when the page states WiFi / internet included
- listing_date
- last_updated
- views_count
- all_images (array of all image URLs found)
- floor_plan_url
- virtual_tour_url
- video_url
- listing_url (current page URL)

For contact/about pages look for:
- company_name, owner_name, founded_year
- address, phone, email, whatsapp
- facebook_url, instagram_url, linkedin_url
- twitter_url, youtube_url
- opening_hours, description
- team_members (array)
- services (array)

RULES:
- Extract EVERY visible data point
- Use null for missing fields
- Arrays for multiple values
- Numbers as numbers (not strings)
- Full URLs for images (not relative)
- Return ONLY valid JSON, no markdown
- AGENT vs OWNER: If both appear, put the estate-agency representative in agent_name/agent_phone/agent_email and the property owner/vendor/landlord in owner_name/owner_phone/owner_email. If only one person is shown, follow the on-page label ("Agent", "Listed by", "Owner", "Vendor"); if still unclear, use agent_* and leave owner_* null.

Page URL: {url}

HTML Content:
{html}
"""


async def smart_scrape(url: str) -> dict:
    """Layered scrape via ScraperEngine (httpx → Playwright → proxy)."""
    r = await _engine_singleton.scrape(url)
    html = r.get("html") or ""
    ok = bool(r.get("success")) and len(html.strip()) > 0
    return {"success": ok, "html": html}


def _normalize_base(url: str) -> str:
    u = url.strip()
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    return u.rstrip("/")


def _norm_netloc(netloc: str) -> str:
    return netloc.lower().replace("www.", "")


def _is_internal(href: str, base_netloc: str) -> bool:
    try:
        p = urlparse(href)
        if p.scheme not in ("http", "https"):
            return False
        hn = _norm_netloc(p.netloc)
        bn = _norm_netloc(base_netloc)
        return hn == bn or hn.endswith("." + bn)
    except Exception:
        return False


def _norm_crawl_key(url: str) -> str:
    """Normalize URL for visit deduplication (fragment stripped; host lowercased)."""
    u = (url or "").split("#")[0].strip()
    if not u:
        return ""
    try:
        p = urlparse(u)
        if p.scheme not in ("http", "https"):
            return ""
        host = (p.netloc or "").lower()
        if host.startswith("www."):
            host = host[4:]
        path = p.path or "/"
        query = p.query
        return f"{p.scheme.lower()}://{host}{path}" + (f"?{query}" if query else "")
    except Exception:
        return u.rstrip("/").lower()


_ASSET_OR_SKIP_SUFFIX = (
    ".jpg",
    ".jpeg",
    ".png",
    ".gif",
    ".webp",
    ".svg",
    ".ico",
    ".pdf",
    ".zip",
    ".rar",
    ".css",
    ".js",
    ".mjs",
    ".map",
    ".woff",
    ".woff2",
    ".ttf",
    ".eot",
    ".mp4",
    ".webm",
    ".mp3",
    ".xml",
)


def _should_enqueue_for_crawl(href: str) -> bool:
    """Skip binaries / assets when enqueueing URLs for the site crawl."""
    if not href or href.strip().lower().startswith("javascript:"):
        return False
    low = href.lower().split("?")[0]
    if any(low.endswith(sfx) for sfx in _ASSET_OR_SKIP_SUFFIX):
        return False
    if "mailto:" in href or "tel:" in href:
        return False
    return True


# When the homepage is a SPA shell or slow to hydrate, try these paths on the same host.
_CRAWL_SEED_PATHS: tuple[str, ...] = (
    "/properties",
    "/property",
    "/property-search",
    "/listings",
    "/for-sale",
    "/for-rent",
    "/buy",
    "/rent",
    "/search",
    "/results",
    "/all-properties",
    "/en/properties",
    "/en/for-sale",
    "/en/for-rent",
    "/our-properties",
    "/listings-for-sale",
    "/properties-for-sale",
    "/property-for-sale",
    "/property-for-rent",
    "/residential",
    "/commercial",
)


def _heuristic_listing_seed_urls(base_url: str) -> list[str]:
    u = (base_url or "").strip().rstrip("/")
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    try:
        p = urlparse(u)
        origin = f"{p.scheme}://{p.netloc}"
    except Exception:
        return []
    out: list[str] = []
    for path in _CRAWL_SEED_PATHS:
        full = urljoin(origin + "/", path)
        if full not in out:
            out.append(full)
    return out


def _ingest_hrefs_into_crawl(
    base_for_join: str,
    href_text_pairs: list[tuple[str, str]],
    merged: dict[str, dict],
    queue: deque[str],
    scheduled: set[str],
    visited_ok: set[str],
    failed_keys: set[str],
    crawl_domain: str,
) -> None:
    """Resolve relative hrefs, store internal links, enqueue for BFS."""
    for raw_href, tx in href_text_pairs:
        href = (raw_href or "").strip()
        if not href:
            continue
        if href.startswith("#"):
            continue
        if not _should_enqueue_for_crawl(href):
            continue
        try:
            abs_url = urljoin(base_for_join.rstrip("/") + "/", href)
        except Exception:
            continue
        if not _is_internal(abs_url, crawl_domain):
            continue
        clean = abs_url.split("#")[0].strip()
        if not clean:
            continue
        _store_link(merged, clean, (tx or "")[:80], False)
        if not _should_enqueue_for_crawl(clean):
            continue
        ck = _norm_crawl_key(clean)
        if (
            ck
            and ck not in visited_ok
            and ck not in failed_keys
            and ck not in scheduled
            and len(queue) < 8000
        ):
            scheduled.add(ck)
            queue.append(clean)


async def _try_dismiss_cookie_banners(page) -> None:
    """Best-effort: sites that hide nav until consent is clicked."""
    labels = (
        "Accept all cookies",
        "Accept all",
        "Accept Cookies",
        "Accept",
        "I Agree",
        "Agree",
        "Allow all",
        "Allow all cookies",
        "OK",
        "Got it",
    )
    for lbl in labels:
        try:
            await page.get_by_role("button", name=re.compile(re.escape(lbl), re.I)).first.click(timeout=900)
            await page.wait_for_timeout(600)
            return
        except Exception:
            continue
    try:
        await page.locator("text=/^\\s*Accept\\s*$/i").first.click(timeout=700)
        await page.wait_for_timeout(500)
    except Exception:
        pass


def _store_link(store: dict[str, dict], url: str, text: str, is_nav: bool) -> None:
    u = (url or "").split("#")[0].strip()
    if not u:
        return
    key = u.rstrip("/") or u
    text = (text or "").strip()[:80]
    if key not in store:
        store[key] = {"url": key, "text": text, "is_nav": bool(is_nav)}
        return
    cur = store[key]
    if len(text) > len(cur.get("text") or ""):
        cur["text"] = text
    cur["is_nav"] = bool(cur.get("is_nav")) or bool(is_nav)


_EXCLUDE_FRAGMENTS = (
    "#",
    "mailto:",
    "tel:",
    "javascript:",
    "facebook.com",
    "twitter.com",
    "instagram.com",
    "linkedin.com",
    "youtube.com",
    "whatsapp.com",
    "/wp-admin",
    "/wp-login",
    "/wp-content/uploads",
    ".pdf",
    ".jpg",
    ".png",
    ".zip",
)

_MUST_EXCLUDE_PATHS = (
    "/service/",
    "/services/",
    "/management",
    "/after-sales",
    "/valuation",
    "/registration",
    "/blog/",
    "/news/",
    "/careers",
    "/privacy",
    "/terms",
    "/cookie",
    "/login",
    "/register",
    "/sitemap",
    "/tag/",
    "/category/",
)

_LISTING_SIGNALS = (
    "/properties",
    "/listings",
    "/for-sale",
    "/for-rent",
    "/buy",
    "/rent",
    "/sale",
    "/search",
    "/results",
    "/all-properties",
)

_PROPERTY_SIGNALS = (
    "/property/",
    "/listing/",
    "/villa/",
    "/apartment/",
    "/penthouse/",
    "/studio/",
    "/house/",
    "/home/",
    "/ref/",
    "/ref=",
)


def _bucket_for_link(url: str, text: str) -> str:
    u = url.lower()
    t = text.lower()

    if any(excl in u for excl in _MUST_EXCLUDE_PATHS):
        return "other_pages"

    if any(sig in u for sig in _LISTING_SIGNALS):
        return "listing_pages"

    if any(sig in u for sig in _PROPERTY_SIGNALS):
        return "property_pages"
    if re.search(r"/[a-z0-9][a-z0-9-]*-\d+/?$", u) or re.search(r"/\d{4,}/?$", u):
        return "property_pages"

    if any(x in u or x in t for x in ("about", "team", "who-we-are", "our-story", "meet")):
        return "about_pages"

    if any(x in u or x in t for x in ("contact", "reach", "get-in-touch", "find-us", "location")):
        return "contact_pages"

    return "other_pages"


def _group_classified_links(
    merged: dict[str, dict],
    website_url: str,
) -> dict[str, list[dict]]:
    grouped = {
        "property_pages": [],
        "listing_pages": [],
        "about_pages": [],
        "contact_pages": [],
        "other_pages": [],
    }
    seen: set[str] = set()
    home = website_url.split("#")[0].strip().rstrip("/")

    for _k, link_obj in merged.items():
        url = link_obj.get("url") or ""
        if not url:
            continue
        ul = url.lower()
        if any(excl in ul for excl in _EXCLUDE_FRAGMENTS):
            continue
        if url in seen:
            continue
        seen.add(url)
        if url.rstrip("/") == home:
            continue

        text = str(link_obj.get("text") or "")

        bucket = _bucket_for_link(url, text)
        grouped[bucket].append(
            {
                "url": url,
                "text": text,
                "is_nav": bool(link_obj.get("is_nav")),
            }
        )

    for key in grouped:
        grouped[key].sort(key=lambda x: x.get("url", ""))

    return grouped


class DiscoverBody(BaseModel):
    city: str
    country: str


class FetchUrlsBody(BaseModel):
    website_url: str
    """Breadth-first Playwright crawl; stops after this many pages successfully loaded."""
    max_pages: int = Field(default=120, ge=1, le=800)


FetchUrlsRequest = FetchUrlsBody


class ExtractBody(BaseModel):
    urls: list[str] = Field(default_factory=list)


class SaveBody(BaseModel):
    data: list[dict] = Field(default_factory=list)
    agency_name: str
    city: str
    country: str
    website_url: str = ""


class ExportExcelBody(BaseModel):
    data: list[dict] = Field(default_factory=list)
    filename: str = "workbench-export"


@router.post("/discover")
async def workbench_discover(body: DiscoverBody):
    """Apify Google Places search for real estate agencies; country is fixed to Malta."""
    city = (body.city or "").strip()
    if not city:
        raise HTTPException(status_code=400, detail="city is required")
    if not (settings.apify_api_token or "").strip():
        raise HTTPException(
            status_code=503,
            detail="APIFY_API_TOKEN is not set on the server. Add it to backend/.env and restart uvicorn.",
        )
    loop = asyncio.get_running_loop()
    try:
        agencies = await loop.run_in_executor(None, lambda: discover_agencies_sync(city, "Malta"))
    except Exception as exc:
        logger.exception("Apify discover failed for city=%s", city)
        raise HTTPException(
            status_code=502,
            detail=f"Apify discovery failed: {exc}",
        ) from exc
    return agencies


@router.post("/fetch-urls")
async def workbench_fetch_urls(request: FetchUrlsRequest):
    """
    Breadth-first Playwright crawl on the same registrable domain: visits up to ``max_pages`` HTML
    pages and collects every internal ``<a href>`` (same behaviour as a shallow site map).
    Falls back to httpx + BeautifulSoup on the homepage only if Playwright fails completely.
    """
    website_url = request.website_url.strip()
    if not website_url.startswith(("http://", "https://")):
        website_url = "https://" + website_url

    parsed = urlparse(website_url)
    autocorrect_note: str | None = None
    # Common typo seen in Malta domains: ".com.m" (missing trailing "t")
    if (parsed.netloc or "").lower().endswith(".com.m"):
        fixed_host = parsed.netloc[:-1] + "t"
        parsed = parsed._replace(netloc=fixed_host)
        website_url = urlunparse(parsed)
        autocorrect_note = f"Input host looked invalid, auto-corrected to {fixed_host}"

    domain = parsed.netloc
    if not domain:
        return {
            "website_url": website_url,
            "total_urls": 0,
            "domain": "",
            "groups": {
                "property_pages": [],
                "listing_pages": [],
                "about_pages": [],
                "contact_pages": [],
                "other_pages": [],
            },
            "error": "Invalid website_url",
        }

    base = website_url.split("#")[0].rstrip("/")
    crawl_domain = domain
    max_pages = min(max(1, request.max_pages), 800)
    merged: dict[str, dict] = {}
    playwright_error: str | None = None
    pages_visited = 0
    discovered_refs: set[str] = set()

    try:
        from playwright.async_api import async_playwright
        from playwright_stealth import stealth_async

        seed = base if urlparse(base).path not in ("", "/") else (base + "/")
        _store_link(merged, seed, "Seed page", True)

        queue: deque[str] = deque([seed])
        scheduled: set[str] = set()
        sk = _norm_crawl_key(seed)
        if sk:
            scheduled.add(sk)
        visited_ok: set[str] = set()
        failed_keys: set[str] = set()

        for seed_extra in _heuristic_listing_seed_urls(base):
            ck0 = _norm_crawl_key(seed_extra)
            if not ck0 or ck0 in scheduled:
                continue
            if not _is_internal(seed_extra, domain):
                continue
            if not _should_enqueue_for_crawl(seed_extra):
                continue
            scheduled.add(ck0)
            queue.append(seed_extra.split("#")[0].strip())
            _store_link(merged, seed_extra.split("#")[0].strip(), "Heuristic listing index", False)

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled"],
            )
            try:
                context = await browser.new_context(
                    viewport={"width": 1280, "height": 800},
                    user_agent=(
                        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                    ),
                    locale="en-US",
                )
                page = await context.new_page()
                await stealth_async(page)

                while queue and pages_visited < max_pages:
                    page_url = queue.popleft()
                    nk = _norm_crawl_key(page_url)
                    if not nk or nk in visited_ok or nk in failed_keys:
                        continue

                    try:
                        await page.goto(page_url, wait_until="load", timeout=32_000)
                        await _try_dismiss_cookie_banners(page)
                        try:
                            await page.wait_for_load_state("networkidle", timeout=5_000)
                        except Exception:
                            pass
                        await page.wait_for_timeout(1_000)
                        try:
                            await page.evaluate(
                                "() => window.scrollTo(0, Math.min(1200, document.body.scrollHeight || 0))"
                            )
                            await page.wait_for_timeout(450)
                        except Exception:
                            pass

                        final_url = (page.url or page_url).strip()
                        final_host = urlparse(final_url).netloc or ""
                        if final_host and _is_internal(final_url, domain):
                            crawl_domain = final_host

                        link_objs = await page.evaluate(
                            """() => {
                              const seen = new Set();
                              const out = [];
                              const push = (href, text) => {
                                if (!href || href.indexOf('http') !== 0) return;
                                if (seen.has(href)) return;
                                seen.add(href);
                                out.push({
                                  href,
                                  text: (text || '').trim().slice(0, 120),
                                });
                              };
                              for (const el of document.querySelectorAll('a[href]')) {
                                push(el.href, el.innerText);
                              }
                              for (const el of document.querySelectorAll(
                                '[data-href], [data-url], [data-link], [data-to]'
                              )) {
                                const raw =
                                  el.getAttribute('data-href') ||
                                  el.getAttribute('data-url') ||
                                  el.getAttribute('data-link') ||
                                  el.getAttribute('data-to');
                                if (!raw || raw.indexOf('javascript:') === 0) continue;
                                try {
                                  const abs = new URL(raw, document.baseURI).href;
                                  push(abs, el.textContent || el.getAttribute('aria-label') || '');
                                } catch (e) {}
                              }
                              return out;
                            }"""
                        )
                    except Exception as nav_exc:
                        logger.debug("Workbench crawl skip %s: %s", page_url, nav_exc)
                        failed_keys.add(nk)
                        continue

                    visited_ok.add(nk)
                    pages_visited += 1

                    for link_obj in link_objs or []:
                        href = (link_obj or {}).get("href") or ""
                        if not href or not _should_enqueue_for_crawl(href):
                            continue
                        if not _is_internal(href, crawl_domain):
                            continue
                        clean = href.split("#")[0].strip()
                        tx = str((link_obj or {}).get("text") or "")
                        _store_link(merged, clean, tx, False)
                        if not _should_enqueue_for_crawl(clean):
                            continue
                        ck = _norm_crawl_key(clean)
                        if (
                            ck
                            and ck not in visited_ok
                            and ck not in failed_keys
                            and ck not in scheduled
                            and len(queue) < 8000
                        ):
                            scheduled.add(ck)
                            queue.append(clean)

                    try:
                        html_doc = await page.content()
                        discovered_refs |= collect_reference_tokens_from_html(html_doc)
                        soup = BeautifulSoup(html_doc, "html.parser")
                        join_base = final_url.split("#")[0].strip() or page_url
                        pairs: list[tuple[str, str]] = []
                        for a in soup.find_all("a", href=True):
                            raw = (a.get("href") or "").strip()
                            if raw:
                                pairs.append((raw, a.get_text(strip=True)[:80]))
                        for ar in soup.select("area[href]"):
                            raw = (ar.get("href") or "").strip()
                            if raw:
                                pairs.append((raw, ""))
                        _ingest_hrefs_into_crawl(
                            join_base,
                            pairs,
                            merged,
                            queue,
                            scheduled,
                            visited_ok,
                            failed_keys,
                            crawl_domain,
                        )
                    except Exception:
                        pass

                    await asyncio.sleep(0.28)

            finally:
                await browser.close()

    except Exception as exc:
        playwright_error = str(exc)
        logger.error("Playwright fetch-urls error: %s", exc)
        try:
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            }
            async with httpx.AsyncClient(timeout=25.0, follow_redirects=True, headers=headers) as client:
                r = await client.get(base + "/")
                if r.status_code != 200:
                    r = await client.get(base)
                if r.status_code == 200 and len(r.text) > 200:
                    discovered_refs |= collect_reference_tokens_from_html(r.text)
                    soup = BeautifulSoup(r.text, "html.parser")
                    for a in soup.find_all("a", href=True):
                        raw = (a.get("href") or "").strip()
                        if not raw:
                            continue
                        href = urljoin(base + "/", raw)
                        if _is_internal(href, crawl_domain):
                            _store_link(merged, href, a.get_text(strip=True)[:80], False)
        except Exception as exc2:
            logger.error("Workbench fetch-urls httpx fallback failed: %s", exc2)

    groups = _group_classified_links(merged, base)
    total = sum(len(v) for v in groups.values())

    all_urls_list = sorted(merged.keys())
    urls_by_reference: dict[str, list[str]] = {}
    for ref in sorted(discovered_refs, key=len, reverse=True):
        hits = _urls_containing_reference(ref, all_urls_list)
        if hits:
            urls_by_reference[ref] = hits

    warning: str | None = None
    if total == 0 and playwright_error:
        if "Executable doesn't exist" in playwright_error or "BrowserType.launch" in playwright_error:
            warning = (
                "Playwright Chromium is not installed. In your project venv run: "
                "python -m playwright install chromium — then restart the API and try again."
            )
        else:
            warning = (
                f"Playwright failed ({playwright_error[:220]}). "
                "HTTP fallback returned no usable links (many sites need a real browser)."
            )
    elif total == 0:
        warning = (
            "No internal links were collected after crawling (same-domain <a> / data-href links). "
            "Try: paste a listings or properties index URL instead of only the homepage; "
            "confirm the site does not block automated browsers; or increase Max pages after "
            "the homepage loads navigation in JavaScript."
        )

    out: dict = {
        "website_url": base,
        "total_urls": total,
        "domain": crawl_domain,
        "groups": groups,
        "pages_visited": pages_visited,
        "crawl_max_pages": max_pages,
        "all_urls": all_urls_list,
        # Reference tokens seen in crawled HTML (grids/cards), mapped to internal URLs that contain that substring.
        "references_from_html": sorted(discovered_refs),
        "urls_by_reference": urls_by_reference,
    }
    if warning:
        out["warning"] = warning
    if autocorrect_note:
        out["autocorrect_note"] = autocorrect_note
    return out


@router.post("/extract")
async def workbench_extract(body: ExtractBody):
    urls = [u.strip() for u in body.urls if u and u.strip()]
    if not urls:
        return {"results": [], "total": 0}

    if not settings.openai_api_key:
        return {
            "results": [
                {
                    "url": u,
                    "success": False,
                    "error": "OPENAI_API_KEY is not configured",
                    "data": None,
                    "kind": None,
                }
                for u in urls
            ],
            "total": 0,
        }

    sem = asyncio.Semaphore(4)
    client = AsyncOpenAI(api_key=settings.openai_api_key)

    async def one(url: str) -> dict:
        async with sem:
            try:
                result = await smart_scrape(url)
                if not result.get("success"):
                    return {
                        "url": url,
                        "success": False,
                        "error": "Empty HTML (scrape failed)",
                        "data": None,
                        "kind": None,
                    }

                html = result["html"]
                soup = BeautifulSoup(html, "html.parser")

                json_ld = ""
                for tag in soup.find_all("script", {"type": "application/ld+json"}):
                    json_ld += (tag.string or "") + "\n"

                meta_data: dict[str, str] = {}
                for meta in soup.find_all("meta"):
                    name = meta.get("name") or meta.get("property") or ""
                    content = meta.get("content") or ""
                    if name and content:
                        meta_data[str(name)] = str(content)

                combined = f"""PAGE URL: {url}

JSON-LD STRUCTURED DATA (most accurate):
{json_ld[:3000] if json_ld else "None"}

META TAGS:
{json.dumps(meta_data, indent=2)[:2000]}

HTML CONTENT:
{html[:15000]}
"""

                prompt = COMPREHENSIVE_EXTRACT_PROMPT.format(url=url, html=combined)

                response = await client.chat.completions.create(
                    model="gpt-4o",
                    messages=[
                        {
                            "role": "system",
                            "content": "You are a precise web data extractor. Return only valid JSON.",
                        },
                        {"role": "user", "content": prompt},
                    ],
                    max_tokens=4000,
                    temperature=0,
                )
                raw = response.choices[0].message.content or ""
                extracted = parse_json_safely(raw)

                if isinstance(extracted, dict) and extracted:
                    extracted["_source_url"] = url
                    extracted["_scraped_at"] = datetime.now(timezone.utc).isoformat()
                    return {
                        "url": url,
                        "success": True,
                        "error": None,
                        "data": extracted,
                        "kind": "comprehensive",
                    }

                return {
                    "url": url,
                    "success": False,
                    "error": "LLM returned empty or invalid JSON",
                    "data": None,
                    "kind": None,
                }
            except Exception as exc:
                logger.exception("Workbench extract failed for %s", url)
                return {"url": url, "success": False, "error": str(exc), "data": None, "kind": None}

    results = await asyncio.gather(*[one(u) for u in urls])
    ok_n = sum(1 for r in results if r.get("success"))
    return {"results": list(results), "total": ok_n}


class UniversalDiscoverPropertiesBody(BaseModel):
    listing_url: str = ""


class UniversalExtractPropertiesBody(BaseModel):
    urls: list[str] = Field(default_factory=list)
    listing_url: str = ""


class UniversalExtractSingleBody(BaseModel):
    url: str = ""


class QualifyPropertyUrlsBody(BaseModel):
    urls: list[str] = Field(default_factory=list, max_length=500)
    require_agent: bool = False
    concurrency: int = Field(default=6, ge=1, le=12)


class MatchReferenceUrlsBody(BaseModel):
    reference: str = ""
    title_hint: str = Field(
        default="",
        description="When reference is missing or weak, match URLs whose path contains title words / slug.",
    )
    urls: list[str] = Field(default_factory=list)
    max_scan: int = Field(default=400, ge=1, le=2000)
    max_matches: int = Field(default=25, ge=1, le=200)
    concurrency: int = Field(default=6, ge=1, le=12)


def _ref_variants(ref: str) -> set[str]:
    r = (ref or "").strip().lower()
    if not r:
        return set()
    out = {r}
    out.add(r.replace(" ", ""))
    out.add(r.replace("-", ""))
    out.add(r.replace("_", ""))
    out.add(r.replace(" ", "-"))
    out.add(r.replace(" ", "_"))
    return {x for x in out if x}


def _title_url_match_signals(title: str) -> list[str]:
    """Words / slug from listing title to match against URL paths (detail pages often embed the name)."""
    t = (title or "").strip().lower()
    if len(t) < 4:
        return []
    parts = [p.strip("-_") for p in re.split(r"[^\w]+", t) if p.strip("-_")]
    words = [p for p in parts if len(p) >= 5]
    slug = "-".join(parts)
    out: list[str] = []
    seen: set[str] = set()
    for x in words:
        if x not in seen:
            seen.add(x)
            out.append(x)
    if len(slug) >= 8 and slug not in seen:
        out.append(slug)
    return out[:14]


@router.post("/match-reference-urls")
async def workbench_match_reference_urls(body: MatchReferenceUrlsBody):
    """
    Find crawl URLs likely belonging to a selected property.

    - With `reference`: match URL text, then scan page HTML for the same token (and common variants).
    - With `title_hint` (listing title): also match URL path segments and HTML text using title words /
      slug, for sites that use SEO URLs without a numeric ref.
    """
    ref = (body.reference or "").strip()
    title_hint = (body.title_hint or "").strip()
    urls = [u.strip() for u in (body.urls or []) if u and str(u).strip()]
    if not urls or (not ref and not title_hint):
        return {"reference": ref, "matched": [], "scanned": 0}

    max_scan = min(len(urls), body.max_scan)
    urls = urls[:max_scan]
    variants = _ref_variants(ref)
    title_signals = _title_url_match_signals(title_hint)

    sem = asyncio.Semaphore(body.concurrency)

    async def one(u: str) -> dict | None:
        ul = u.lower()
        if variants and any(v in ul for v in variants):
            return {"url": u, "source": "url"}
        if title_signals:
            try:
                pth = unquote(urlparse(u).path).lower()
            except Exception:
                pth = (urlparse(u).path or "").lower()
            for sig in title_signals:
                if len(sig) >= 6 and sig in pth:
                    return {"url": u, "source": "title_url"}
        if not variants and not title_signals:
            return None
        async with sem:
            try:
                r = await _engine_singleton.scrape(u)
                html = (r.get("html") or "").lower()
            except Exception:
                return None
        if not html:
            return None
        # Light normalization to catch variants in text/markup.
        h_compact = html.replace(" ", "").replace("-", "").replace("_", "")
        if variants:
            for v in variants:
                if v in html or v.replace(" ", "") in h_compact or v.replace("-", "") in h_compact:
                    return {"url": u, "source": "html"}
        if title_signals:
            for sig in title_signals:
                if len(sig) >= 6 and sig in html:
                    return {"url": u, "source": "html_title"}
        return None

    rows = await asyncio.gather(*[one(u) for u in urls])
    matched = [r for r in rows if r]
    # Prefer direct URL matches first.
    def _rank(m: dict) -> tuple[int, int]:
        src = (m or {}).get("source") or ""
        order = {"url": 0, "title_url": 1, "html": 2, "html_title": 3}
        return (order.get(src, 9), 0)

    matched.sort(key=_rank)
    matched = matched[: body.max_matches]
    return {"reference": ref, "matched": matched, "scanned": max_scan}


@router.post("/qualify-property-urls")
async def workbench_qualify_property_urls(body: QualifyPropertyUrlsBody):
    """
    Quick HTML/JSON-LD scan per URL (ScraperEngine) — keep URLs that look like property detail pages
    (reference, contact, bed/bath/area) before expensive LLM extract-single.
    """
    from backend.scraper.property_url_qualifier import qualify_urls_batch

    urls = [u.strip() for u in (body.urls or []) if u and str(u).strip()][:500]
    if not urls:
        return {"qualified_total": 0, "rejected_total": 0, "qualified": [], "rejected_sample": []}

    qualified, rejected = await qualify_urls_batch(
        urls,
        _engine_singleton.scrape,
        concurrency=body.concurrency,
        require_agent=body.require_agent,
    )
    return {
        "qualified_total": len(qualified),
        "rejected_total": len(rejected),
        "qualified": qualified,
        "rejected_sample": rejected[:40],
    }


@router.post("/discover-properties")
async def workbench_discover_properties(body: UniversalDiscoverPropertiesBody):
    from backend.scraper.universal_extractor import extract_property_urls_from_listing

    listing_url = (body.listing_url or "").strip()
    if not listing_url:
        return {"total": 0, "properties": [], "error": "listing_url is required"}
    # Prevent runaway multi-page Playwright crawls from hanging the UI indefinitely.
    discover_timeout_sec = 720.0
    try:
        return await asyncio.wait_for(extract_property_urls_from_listing(listing_url), timeout=discover_timeout_sec)
    except asyncio.TimeoutError:
        logger.warning("discover-properties timed out after %ss for %s", discover_timeout_sec, listing_url)
        return {
            "total": 0,
            "properties": [],
            "listing_pages_scanned": None,
            "error": (
                "Discovery timed out (12 min limit). Paste your agency’s property search / listings URL "
                "(not only the homepage), or fewer pages may finish faster."
            ),
        }


@router.post("/extract-properties")
async def workbench_extract_properties(body: UniversalExtractPropertiesBody):
    from backend.scraper.universal_extractor import extract_property_detail_universal

    urls = [u.strip() for u in (body.urls or []) if u and str(u).strip()]
    if not urls:
        return {"results": [], "total": 0}

    if not settings.openai_api_key:
        return {
            "results": [
                {"url": u, "success": False, "error": "OPENAI_API_KEY is not configured", "result": None}
                for u in urls
            ],
            "total": 0,
        }

    results: list[dict] = []
    seen_ref: set[str] = set()

    def _meaningful_row(d: dict) -> bool:
        return any(
            d.get(k) is not None and str(d.get(k)).strip() != ""
            for k in ("title", "reference_number", "price", "bedrooms", "internal_sqm", "total_sqm")
        )

    for u in urls:
        try:
            data = await extract_property_detail_universal(u, take_screenshot=False)
            if not isinstance(data, dict):
                results.append({"url": u, "success": False, "error": "Invalid response", "result": data})
                continue
            if data.get("error") is not None and not _meaningful_row(data):
                results.append(
                    {"url": u, "success": False, "error": str(data.get("error")), "result": data},
                )
                continue
            ref = str(data.get("reference_number") or data.get("reference") or "").strip()
            key = ref.lower() if ref else u
            if key in seen_ref:
                continue
            seen_ref.add(key)
            results.append({"url": u, "success": True, "error": None, "result": data})
        except Exception as exc:
            logger.exception("extract-properties failed for %s", u)
            results.append({"url": u, "success": False, "error": str(exc), "result": None})

    ok = sum(1 for r in results if r.get("success"))
    return {"results": results, "total": ok}


@router.post("/extract-single")
async def workbench_extract_single(body: UniversalExtractSingleBody):
    from backend.scraper.universal_extractor import extract_property_detail_universal

    u = (body.url or "").strip()
    if not u:
        return {"result": None, "url": "", "error": "url is required"}
    data = await extract_property_detail_universal(u, take_screenshot=True)
    return {"result": data, "url": u}


@router.post("/save")
async def workbench_save(body: SaveBody):
    from backend.routers.scraper import _build_property_row

    website = (body.website_url or "").strip()
    if not website.startswith("http"):
        website = "https://" + website.lstrip("/")

    saved_props = 0
    _, session_factory = _get_engine()
    async with session_factory() as db:
        agency_payload = {
            "name": body.agency_name.strip() or "Workbench Import",
            "website_url": website,
            "city": body.city,
            "country": body.country,
            "scrape_level": 2,
            "scrape_status": "done",
        }
        agency = await crud.upsert_agency(db, agency_payload)
        if not agency:
            return {"saved": 0, "error": "Could not save agency"}

        for row in body.data:
            if not isinstance(row, dict):
                continue
            inner = row.get("data") if isinstance(row.get("data"), dict) else row
            if not isinstance(inner, dict) or not inner.get("title"):
                continue
            pr = _build_property_row(inner, agency.id, body.city, body.country)
            if pr.get("title"):
                await crud.create_property(db, pr)
                saved_props += 1

    return {"saved": saved_props}


def _flatten_row(obj: dict, prefix: str = "") -> dict[str, object]:
    flat: dict[str, object] = {}
    for k, v in obj.items():
        key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict):
            flat.update(_flatten_row(v, key + "."))
        elif isinstance(v, list):
            flat[key] = ", ".join(str(x) for x in v[:80])
        else:
            flat[key] = v
    return flat


@router.post("/export-excel")
async def workbench_export_excel(body: ExportExcelBody):
    rows_in = body.data or []
    if not rows_in:
        buf = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["(empty)"])
        wb.save(buf)
        buf.seek(0)
        name = re.sub(r"[^\w.\-]", "_", body.filename or "export") + ".xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{name}"'},
        )

    flat_rows = []
    for r in rows_in:
        if not isinstance(r, dict):
            continue
        data = r.get("data") if isinstance(r.get("data"), dict) else r
        base = {"source_url": r.get("url", "")}
        if isinstance(data, dict):
            merged = {**base, **_flatten_row(data)}
        else:
            merged = dict(base)
        flat_rows.append(merged)

    keys: list[str] = []
    seen: set[str] = set()
    for fr in flat_rows:
        for k in fr.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    header_font = Font(bold=True)
    for col, key in enumerate(keys, start=1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = header_font
    for ri, fr in enumerate(flat_rows, start=2):
        for ci, key in enumerate(keys, start=1):
            ws.cell(row=ri, column=ci, value=fr.get(key))

    ws.freeze_panes = "A2"
    for ci, key in enumerate(keys, start=1):
        max_len = len(str(key))
        for row in ws.iter_rows(min_col=ci, max_col=ci, min_row=1, max_row=min(ws.max_row, 500)):
            for c in row:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = re.sub(r"[^\w.\-]", "_", body.filename or "workbench-export") + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )

